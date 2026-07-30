from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from trade_helper.ledger import (
    AccountSnapshot,
    CashFlow,
    Ledger,
    PositionSnapshot,
    Trade,
)


ASSETS = {
    "SP500": "513500",
    "NASDAQ": "513100",
    "DIVIDEND": "515450",
}
SOURCES = {"MANUAL", "EXCEL_IMPORT", "APP_FORM", "BROKER_CSV"}
TRADE_STATUSES = {
    "SUBMITTED", "PARTIALLY_FILLED", "FILLED", "CANCELLED", "REJECTED"
}
CASH_FLOW_TYPES = {
    "DEPOSIT", "WITHDRAWAL", "DIVIDEND", "INTEREST", "FEE", "TAX", "ADJUSTMENT"
}
SHANGHAI = ZoneInfo("Asia/Shanghai")


@dataclass(frozen=True)
class ImportIssue:
    sheet: str
    row: int
    field: str
    message: str


@dataclass(frozen=True)
class ImportPreview:
    source_name: str
    content_hash: str
    snapshots: tuple[tuple[AccountSnapshot, tuple[PositionSnapshot, ...]], ...]
    trades: tuple[Trade, ...]
    cash_flows: tuple[CashFlow, ...]
    issues: tuple[ImportIssue, ...]

    @property
    def valid(self) -> bool:
        return not self.issues

    @property
    def row_counts(self) -> dict[str, int]:
        return {
            "snapshots": len(self.snapshots),
            "positions": sum(len(item[1]) for item in self.snapshots),
            "trades": len(self.trades),
            "cash_flows": len(self.cash_flows),
        }


class InvalidImport(ValueError):
    def __init__(self, issues: tuple[ImportIssue, ...]) -> None:
        super().__init__(f"Excel import has {len(issues)} validation error(s)")
        self.issues = issues


def _rows(sheet) -> list[tuple[int, dict[str, object]]]:
    headers = [cell.value for cell in sheet[3]]
    result = []
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        if all(value is None or value == "" for value in values):
            continue
        result.append((row_number, dict(zip(headers, values))))
    return result


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: object, *, positive: bool = False) -> Decimal:
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError) as error:
        raise ValueError("must be a number") from error
    if positive and parsed <= 0:
        raise ValueError("must be greater than zero")
    return parsed


def _integer(value: object, *, positive: bool = False) -> int:
    parsed = _decimal(value)
    if parsed != parsed.to_integral_value():
        raise ValueError("must be an integer")
    result = int(parsed)
    if (positive and result <= 0) or (not positive and result < 0):
        raise ValueError("must be positive" if positive else "must be non-negative")
    return result


def _datetime(value: object) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.strip())
        except ValueError as error:
            raise ValueError("must be an ISO date/time") from error
    else:
        raise ValueError("must be a date/time")
    return parsed.replace(tzinfo=SHANGHAI) if parsed.tzinfo is None else parsed


def _validate_asset(asset_id: str, etf_code: str) -> None:
    if asset_id not in ASSETS:
        raise ValueError(f"unsupported asset_id: {asset_id}")
    if etf_code != ASSETS[asset_id]:
        raise ValueError(f"{asset_id} must use ETF {ASSETS[asset_id]}")


def preview_workbook(path: str | Path) -> ImportPreview:
    source = Path(path)
    content = source.read_bytes()
    digest = hashlib.sha256(content).hexdigest()
    workbook = load_workbook(source, data_only=False, read_only=True)
    required = {"账户快照", "持仓快照", "交易流水", "资金流水"}
    issues: list[ImportIssue] = []
    missing = required.difference(workbook.sheetnames)
    for name in sorted(missing):
        issues.append(ImportIssue(name, 0, "", "missing required worksheet"))
    if missing:
        workbook.close()
        return ImportPreview(source.name, digest, (), (), (), tuple(issues))

    snapshot_rows = _rows(workbook["账户快照"])
    position_rows = _rows(workbook["持仓快照"])
    trade_rows = _rows(workbook["交易流水"])
    flow_rows = _rows(workbook["资金流水"])
    snapshots: dict[str, AccountSnapshot] = {}
    positions: dict[str, list[PositionSnapshot]] = {}
    trades: list[Trade] = []
    flows: list[CashFlow] = []

    def issue(sheet: str, row: int, field: str, error: Exception | str) -> None:
        issues.append(ImportIssue(sheet, row, field, str(error)))

    for row_number, row in snapshot_rows:
        try:
            snapshot_id = _text(row.get("snapshot_id"))
            if not snapshot_id:
                raise ValueError("snapshot_id is required")
            if snapshot_id in snapshots:
                raise ValueError("duplicate snapshot_id in workbook")
            source_value = _text(row.get("source")) or "EXCEL_IMPORT"
            if source_value not in SOURCES:
                raise ValueError("unsupported source")
            snapshots[snapshot_id] = AccountSnapshot(
                snapshot_id=snapshot_id,
                as_of=_datetime(row.get("as_of")),
                total_assets_cny=_decimal(row.get("total_assets")),
                available_cash_cny=_decimal(row.get("available_cash")),
                frozen_cash_cny=_decimal(row.get("frozen_cash") or 0),
                source="EXCEL_IMPORT",
                notes=_text(row.get("notes")),
            )
            if min(
                snapshots[snapshot_id].total_assets_cny,
                snapshots[snapshot_id].available_cash_cny,
                snapshots[snapshot_id].frozen_cash_cny,
            ) < 0:
                raise ValueError("money values must be non-negative")
        except ValueError as error:
            issue("账户快照", row_number, "", error)

    seen_positions: set[tuple[str, str]] = set()
    for row_number, row in position_rows:
        try:
            snapshot_id = _text(row.get("snapshot_id"))
            asset_id = _text(row.get("asset_id"))
            etf_code = _text(row.get("etf_code")).zfill(6)
            if snapshot_id not in snapshots:
                raise ValueError("snapshot_id does not reference an account snapshot")
            _validate_asset(asset_id, etf_code)
            key = (snapshot_id, asset_id)
            if key in seen_positions:
                raise ValueError("duplicate asset in snapshot")
            seen_positions.add(key)
            market_value = row.get("broker_market_value")
            positions.setdefault(snapshot_id, []).append(
                PositionSnapshot(
                    snapshot_id, asset_id, etf_code,
                    _integer(row.get("quantity")),
                    _decimal(market_value) if market_value not in (None, "") else None,
                    "EXCEL_IMPORT",
                )
            )
        except ValueError as error:
            issue("持仓快照", row_number, "", error)

    seen_trades: set[str] = set()
    for row_number, row in trade_rows:
        try:
            trade_id = _text(row.get("trade_id"))
            asset_id = _text(row.get("asset_id"))
            etf_code = _text(row.get("etf_code")).zfill(6)
            side = _text(row.get("side"))
            status = _text(row.get("status")) or "FILLED"
            if not trade_id or trade_id in seen_trades:
                raise ValueError("trade_id is required and must be unique")
            seen_trades.add(trade_id)
            _validate_asset(asset_id, etf_code)
            if side not in {"BUY", "SELL"}:
                raise ValueError("side must be BUY or SELL")
            if status not in TRADE_STATUSES:
                raise ValueError("unsupported trade status")
            trades.append(
                Trade(
                    trade_id, _datetime(row.get("trade_time")), asset_id, etf_code,
                    side, _integer(row.get("quantity"), positive=True),
                    _decimal(row.get("price"), positive=True), status,
                    "EXCEL_IMPORT", _text(row.get("notes")),
                )
            )
        except ValueError as error:
            issue("交易流水", row_number, "", error)

    seen_flows: set[str] = set()
    for row_number, row in flow_rows:
        try:
            flow_id = _text(row.get("flow_id"))
            flow_type = _text(row.get("flow_type"))
            if not flow_id or flow_id in seen_flows:
                raise ValueError("flow_id is required and must be unique")
            seen_flows.add(flow_id)
            if flow_type not in CASH_FLOW_TYPES:
                raise ValueError("unsupported flow_type")
            amount = _decimal(row.get("amount"))
            if amount == 0:
                raise ValueError("amount must not be zero")
            if flow_type in {"WITHDRAWAL", "FEE", "TAX"} and amount > 0:
                raise ValueError(f"{flow_type} amount must be negative")
            flows.append(
                CashFlow(
                    flow_id, _datetime(row.get("flow_time")), flow_type, amount,
                    _text(row.get("asset_id")) or None, _text(row.get("notes")),
                )
            )
        except ValueError as error:
            issue("资金流水", row_number, "", error)

    workbook.close()
    grouped = tuple(
        (snapshot, tuple(positions.get(snapshot_id, [])))
        for snapshot_id, snapshot in snapshots.items()
    )
    return ImportPreview(
        source.name, digest, grouped, tuple(trades), tuple(flows), tuple(issues)
    )


def commit_preview(ledger: Ledger, preview: ImportPreview) -> bool:
    if not preview.valid:
        raise InvalidImport(preview.issues)
    return ledger.apply_import_batch(
        content_hash=preview.content_hash,
        source_name=preview.source_name,
        snapshots=preview.snapshots,
        trades=preview.trades,
        cash_flows=preview.cash_flows,
    )
