from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import TestCase

from openpyxl import Workbook, load_workbook

from trade_helper.excel_import import InvalidImport, commit_preview, preview_workbook
from trade_helper.ledger import Ledger


HEADERS = {
    "账户快照": [
        "snapshot_id", "as_of", "total_assets", "available_cash",
        "frozen_cash", "source", "notes",
    ],
    "持仓快照": [
        "snapshot_id", "as_of", "asset_id", "etf_code", "quantity",
        "broker_market_value", "source", "notes",
    ],
    "交易流水": [
        "trade_id", "trade_time", "asset_id", "etf_code", "side", "quantity",
        "price", "gross_amount", "net_cash_flow", "status", "order_id",
        "source", "notes",
    ],
    "资金流水": [
        "flow_id", "flow_time", "flow_type", "amount", "asset_id",
        "etf_code", "description", "source", "notes",
    ],
}


class ExcelImportTests(TestCase):
    def setUp(self) -> None:
        self.directory = TemporaryDirectory()
        self.addCleanup(self.directory.cleanup)
        self.root = Path(self.directory.name)
        self.database = Ledger(self.root / "account.db")
        self.database.initialize()

    def make_workbook(self, *, invalid_trade: bool = False) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name, headers in HEADERS.items():
            sheet = workbook.create_sheet(name)
            sheet.append(["title"])
            sheet.append([])
            sheet.append(headers)
        workbook["账户快照"].append(
            ["SNAP-1", "2026-07-30 14:00:00", 500000, 350000, 0, "MANUAL", ""]
        )
        workbook["持仓快照"].append(
            ["SNAP-1", "2026-07-30 14:00:00", "SP500", "513500", 24000, 60000, "MANUAL", ""]
        )
        workbook["交易流水"].append(
            [
                "TRD-1", "2026-07-30 14:10:00", "SP500", "513500",
                "BUY", 0 if invalid_trade else 100, 2.5, None, None,
                "FILLED", "", "MANUAL", "",
            ]
        )
        workbook["资金流水"].append(
            ["FLOW-1", "2026-07-01 09:00:00", "DEPOSIT", 16000, "", "", "", "MANUAL", ""]
        )
        path = self.root / ("invalid.xlsx" if invalid_trade else "valid.xlsx")
        workbook.save(path)
        workbook.close()
        return path

    def test_preview_and_atomic_commit(self) -> None:
        preview = preview_workbook(self.make_workbook())

        self.assertTrue(preview.valid)
        self.assertEqual(
            {"snapshots": 1, "positions": 1, "trades": 1, "cash_flows": 1},
            preview.row_counts,
        )
        self.assertTrue(commit_preview(self.database, preview))
        self.assertEqual(1, self.database.count("account_snapshots"))
        self.assertEqual(1, self.database.count("position_snapshots"))
        self.assertEqual(1, self.database.count("trades"))
        self.assertEqual(1, self.database.count("cash_flows"))

    def test_same_file_is_idempotent(self) -> None:
        preview = preview_workbook(self.make_workbook())
        self.assertTrue(commit_preview(self.database, preview))
        self.assertFalse(commit_preview(self.database, preview))
        self.assertEqual(1, self.database.count("import_batches"))
        self.assertEqual(1, self.database.count("trades"))

    def test_invalid_row_blocks_the_whole_batch(self) -> None:
        preview = preview_workbook(self.make_workbook(invalid_trade=True))

        self.assertFalse(preview.valid)
        with self.assertRaises(InvalidImport):
            commit_preview(self.database, preview)

        self.assertEqual(0, self.database.count("account_snapshots"))
        self.assertEqual(0, self.database.count("cash_flows"))

    def test_formula_only_template_rows_are_ignored(self) -> None:
        path = self.make_workbook()
        workbook = load_workbook(path)
        workbook["交易流水"]["H5"] = "=F5*G5"
        workbook["交易流水"]["I5"] = '=IF(E5="BUY",-H5,H5)'
        workbook.save(path)
        workbook.close()

        preview = preview_workbook(path)

        self.assertTrue(preview.valid)
        self.assertEqual(1, preview.row_counts["trades"])
